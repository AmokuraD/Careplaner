# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import os.path as op
'''
class ExcelPipeline(object):
    def __init__(self):
        try:
            # 检查是否存在文件
            self.wb = load_workbook('./Position.xlsx')
             # 激活工作表
            self.ws = self.wb.active
        except:
             # 实例化
            self.wb = Workbook()
            # 激活工作表
            self.ws = self.wb.active
            self.ws.append(['FirstCategory', 'SecondCategory', 'PositionName', 'PositionDetail'])
            self.wb.save('./Position.xlsx')
            print('Succeed in making xlsx')
        
    def process_item(self, item, spider):
        datas = dict(item)
        line_datas = [''.join(datas['first_ctg']), 
                            ''.join(datas['second_ctg']), 
                            ''.join(datas['positionname']), 
                            ''.join(datas['positiondetail'])]
        self.ws.append(line_datas)
        self.wb.save('./Position.xlsx')
        spider.crawler.stats.inc_value('Success_Append_Excel')
        return item
'''
class ExcelPipeline(object):        
        
    def process_item(self, item, spider):
        datas = dict(item)
        # 根据item的一级类型和二级类型创建excel表单
        dirname = "./{}".format(datas['first_ctg'][0])
        filename = dirname + "/{}.xlsx".format(datas['second_ctg'][0])
        if not op.exists(dirname) :
            os.mkdir(dirname)

        try:
            # 检查是否存在文件
            self.wb = load_workbook(filename)
             # 激活工作表
            self.ws = self.wb.active
        except:
             # 实例化
            self.wb = Workbook()
            # 激活工作表
            self.ws = self.wb.active
            self.ws.append(['FirstCategory', 'SecondCategory', 'PositionName', 'PositionDetail'])
            self.wb.save(filename)
            print('Succeed in making xlsx')
        else:
            line_datas = [datas['first_ctg'][0], 
                          datas['second_ctg'][0], 
                          datas['positionname'][0], 
                          datas['positiondetail'][0],]
            self.ws.append(line_datas)
            self.wb.save(filename)
            spider.crawler.stats.inc_value('Success_Append_Excel')
            return item
